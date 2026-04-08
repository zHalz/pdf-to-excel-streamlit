import pandas as pd
from io import BytesIO
from openpyxl import load_workbook


def exportar_para_excel_completo(df_consolidado, pivot_completa, analise_plano, df_totvs):

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_consolidado.to_excel(writer, sheet_name="Detalhamento", index=False)
        pivot_completa.to_excel(writer, sheet_name="Pivot_Eventos", index=False)
        analise_plano.to_excel(writer, sheet_name="Analise_Plano_Saude", index=False)
        df_totvs.to_excel(writer, sheet_name="Base_TOTVS", index=False)

    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb["Base_TOTVS"]
    ultima_linha = ws.max_row

    linha_titular = ultima_linha + 2
    linha_dependente = ultima_linha + 3

    ws[f"D{linha_titular}"] = "TITULAR"
    ws[f"D{linha_dependente}"] = "DEPENDENTE"

    ws[f"E{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(E2,ROW(E2:E{ultima_linha})-ROW(E2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"F{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(F2,ROW(F2:F{ultima_linha})-ROW(F2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"G{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(G2,ROW(G2:G{ultima_linha})-ROW(G2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'
    ws[f"H{linha_titular}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(H2,ROW(H2:H{ultima_linha})-ROW(H2),0)),(C2:C{ultima_linha}=D{linha_titular})+0)'

    ws[f"E{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(E2,ROW(E2:E{ultima_linha})-ROW(E2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"F{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(F2,ROW(F2:F{ultima_linha})-ROW(F2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"G{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(G2,ROW(G2:G{ultima_linha})-ROW(G2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'
    ws[f"H{linha_dependente}"] = f'=SUMPRODUCT(SUBTOTAL(9,OFFSET(H2,ROW(H2:H{ultima_linha})-ROW(H2),0)),(C2:C{ultima_linha}=D{linha_dependente})+0)'

    final = BytesIO()
    wb.save(final)
    final.seek(0)

    return final